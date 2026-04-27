#!/usr/bin/env python3
"""
Excel ファイルの全構造を解析し、JSON 形式で出力するスクリプト。
対応形式: .xlsx, .xlsm

ディレクトリ規約:
  - Excel ファイルは <project_root>/xlms/ に格納
  - 出力 MD パスは <project_root>/docs/<stem>_仕様書.md に解決される

Usage:
    python3 extract_excel.py <filepath>
"""

import sys
import json
import zipfile
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl が未インストールです。pip3 install openpyxl を実行してください。", file=sys.stderr)
    sys.exit(1)


def extract_sheets_info(wb):
    """シート一覧と基本情報を取得"""
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    return sheets


def extract_cells(wb):
    """全セルの値を取得（data_only=True で読み込んだ wb を使用）"""
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheet_cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val = cell.value
                    if hasattr(val, 'isoformat'):
                        val = val.isoformat()
                    sheet_cells.append({
                        "coord": cell.coordinate,
                        "col_letter": cell.column_letter,
                        "row": cell.row,
                        "value": str(val),
                    })
        if sheet_cells:
            result[name] = sheet_cells
    return result


def extract_formulas(filepath):
    """全数式を取得（data_only=False で読み込み）"""
    wb = openpyxl.load_workbook(filepath, keep_vba=True, data_only=False)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    formulas.append({
                        "coord": cell.coordinate,
                        "formula": val,
                    })
        if formulas:
            result[name] = formulas
    return result


def extract_merged_cells(wb):
    """マージセル範囲を取得"""
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        merged = [str(mc) for mc in ws.merged_cells.ranges]
        if merged:
            result[name] = merged
    return result


def extract_defined_names(wb):
    """名前付き範囲を取得"""
    names = []
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        names.append({
            "name": name,
            "value": dn.attr_text,
        })
    return names


def extract_buttons(filepath):
    """ZIP 内のドローイング・VML からボタンとマクロ割り当てを抽出"""
    buttons = []
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            for fname in z.namelist():
                if 'vmlDrawing' in fname and fname.endswith('.vml'):
                    content = z.read(fname).decode('utf-8', errors='replace')
                    macros = re.findall(r'<x:FmlaMacro>(.*?)</x:FmlaMacro>', content)
                    texts = re.findall(
                        r'<div[^>]*><font[^>]*>(.*?)</font></div>',
                        content, re.DOTALL
                    )
                    for i, macro in enumerate(macros):
                        btn_text = texts[i].strip() if i < len(texts) else "不明"
                        buttons.append({
                            "source_file": fname,
                            "button_text": btn_text,
                            "macro": macro,
                        })
    except Exception as e:
        buttons.append({"error": str(e)})
    return buttons


def extract_external_links(filepath):
    """外部リンク情報を抽出"""
    links = []
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            for fname in z.namelist():
                if 'externalLinks/' in fname and fname.endswith('.xml.rels'):
                    content = z.read(fname).decode('utf-8', errors='replace')
                    targets = re.findall(r'Target="([^"]+)"', content)
                    for t in targets:
                        links.append({"target": t})
                elif 'externalLinks/' in fname and not fname.endswith('.rels'):
                    content = z.read(fname).decode('utf-8', errors='replace')
                    sheet_names = re.findall(r'sheetName val="([^"]+)"', content)
                    if sheet_names:
                        links.append({"sheets_in_external": sheet_names})
    except Exception as e:
        links.append({"error": str(e)})
    return links


def extract_zip_contents(filepath):
    """ZIP 内のファイル一覧"""
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            return z.namelist()
    except Exception:
        return []


def resolve_paths(filepath: str):
    """
    プロジェクトルートを自動探索し、xlms/ と docs/ のパスを解決する。

    優先順位:
      1. 指定ファイルの親が xlms/ なら、そのルートを使用
      2. 指定ファイルの親ディレクトリ内に xlms/ があればそこをルートとして使用
      3. いずれも見つからなければ指定ファイルと同じディレクトリを xlms/、
         同階層の docs/ を出力先とする（なければ作成）
    """
    src = Path(filepath).resolve()
    stem = src.stem

    # ケース1: 既に xlms/ の中にある
    if src.parent.name == "xlms":
        project_root = src.parent.parent
    # ケース2: 親ディレクトリに xlms/ がある
    elif (src.parent / "xlms").is_dir():
        project_root = src.parent
    else:
        project_root = src.parent

    xlms_dir = project_root / "xlms"
    docs_dir = project_root / "docs"

    xlms_dir.mkdir(exist_ok=True)
    docs_dir.mkdir(exist_ok=True)

    # xlms/ 外のファイルは xlms/ にコピー
    dest_excel = xlms_dir / src.name
    if src != dest_excel:
        import shutil
        shutil.copy2(str(src), str(dest_excel))
        print(f"INFO: Excel ファイルを xlms/ にコピーしました: {dest_excel}", file=sys.stderr)

    output_md = docs_dir / f"{stem}_仕様書.md"

    return str(dest_excel), str(output_md), str(project_root)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_excel.py <filepath>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f"ERROR: ファイルが見つかりません: {filepath}", file=sys.stderr)
        sys.exit(1)

    excel_path, output_md_path, project_root = resolve_paths(filepath)

    ext = Path(excel_path).suffix.lower()
    has_vba = ext in ('.xlsm', '.xlsb')

    wb = openpyxl.load_workbook(excel_path, keep_vba=has_vba, data_only=True)

    output = {
        "filepath": excel_path,
        "output_md_path": output_md_path,
        "project_root": project_root,
        "extension": ext,
        "has_vba": has_vba,
        "sheets": extract_sheets_info(wb),
        "cells": extract_cells(wb),
        "formulas": extract_formulas(excel_path),
        "merged_cells": extract_merged_cells(wb),
        "defined_names": extract_defined_names(wb),
        "buttons": extract_buttons(excel_path),
        "external_links": extract_external_links(excel_path),
        "zip_contents": extract_zip_contents(excel_path),
    }

    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
